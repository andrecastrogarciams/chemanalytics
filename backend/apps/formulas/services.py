import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from django.db.models import Max
from django.db import transaction
from openpyxl import load_workbook

from .models import Formula, FormulaItem, FormulaVersion


def parse_date(value: str):
    if hasattr(value, "date"):
        return value.date()
    return datetime.strptime(value, "%Y-%m-%d").date()


def _empty_report():
    return {
        "source_file": None,
        "source_type": None,
        "formulas_created": 0,
        "versions_created": 0,
        "items_created": 0,
        "incomplete_items_created": 0,
        "rejected_rows": [],
    }


def _load_rows_from_csv(path: Path, report):
    grouped = defaultdict(list)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for line_number, row in enumerate(reader, start=2):
            try:
                codpro = row["codpro"].strip()
                codder = row["codder"].strip()
                start_date = parse_date(row["start_date"].strip())
                grouped[(codpro, codder, start_date)].append((line_number, row))
            except Exception as exc:
                report["rejected_rows"].append({"line": line_number, "error": str(exc)})
    return grouped


def _load_rows_from_xlsx(path: Path, report):
    grouped = defaultdict(list)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    for line_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            normalized = {
                "codpro": str(row[0]).strip(),
                "codder": str(row[1]).strip(),
                "chemical_description": str(row[2]).strip(),
                "chemical_code": str(row[3]).strip(),
                "percentual": "" if row[4] is None else str(row[4]).strip(),
                "start_date": row[5],
                "tolerance_pct": "0",
            }
            start_date = parse_date(normalized["start_date"])
            grouped[(normalized["codpro"], normalized["codder"], start_date)].append((line_number, normalized))
        except Exception as exc:
            report["rejected_rows"].append({"line": line_number, "error": str(exc)})
    return grouped


@transaction.atomic
def bootstrap_formulas_from_file(file_path: str):
    path = Path(file_path)
    report = _empty_report()
    report["source_file"] = str(path)
    report["source_type"] = path.suffix.lower().lstrip(".")

    if path.suffix.lower() == ".csv":
        grouped = _load_rows_from_csv(path, report)
    elif path.suffix.lower() == ".xlsx":
        grouped = _load_rows_from_xlsx(path, report)
    else:
        raise ValueError(f"Unsupported file type: {path.suffix}")

    for (codpro, codder, start_date), rows in grouped.items():
        formula, created = Formula.objects.get_or_create(codpro=codpro, codder=codder)
        if created:
            report["formulas_created"] += 1

        existing_version = FormulaVersion.objects.filter(formula=formula, start_date=start_date).first()
        if existing_version:
            version = existing_version
        else:
            next_version = (formula.versions.aggregate(models_max=Max("version_number"))["models_max"] or 0) + 1
            version = FormulaVersion.objects.create(
                formula=formula,
                version_number=next_version,
                start_date=start_date,
                observation="Bootstrap import",
            )
            report["versions_created"] += 1

        seen_codes = set()
        for line_number, row in rows:
            chemical_code = row.get("chemical_code", "").strip()
            if chemical_code in seen_codes:
                report["rejected_rows"].append(
                    {"line": line_number, "error": f"Duplicate chemical_code {chemical_code} in same version"}
                )
                continue
            seen_codes.add(chemical_code)

            try:
                percentual_raw = row["percentual"].strip() if row.get("percentual") is not None else ""
                is_incomplete = percentual_raw == ""
                _, item_created = FormulaItem.objects.get_or_create(
                    formula_version=version,
                    chemical_code=chemical_code,
                    defaults={
                        "chemical_description": row["chemical_description"].strip(),
                        "percentual": None if is_incomplete else percentual_raw,
                        "tolerance_pct": row.get("tolerance_pct", "0").strip() or "0",
                        "is_incomplete": is_incomplete,
                        "incomplete_reason": "Missing percentual in source file" if is_incomplete else None,
                    },
                )
                if item_created:
                    report["items_created"] += 1
                    if is_incomplete:
                        report["incomplete_items_created"] += 1
            except Exception as exc:
                report["rejected_rows"].append({"line": line_number, "error": str(exc)})

    return report
